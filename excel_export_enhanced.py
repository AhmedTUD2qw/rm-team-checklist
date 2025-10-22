#!/usr/bin/env python3
"""
Enhanced Excel export with Cloudinary support and better formatting
"""

import os
import tempfile
import requests
from io import BytesIO
from PIL import Image as PILImage
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from cloudinary_config import upload_excel_to_cloudinary, cleanup_temp_file, is_cloudinary_configured

def download_image_from_cloudinary(image_url, max_size=(200, 200)):
    """
    تحميل صورة من Cloudinary وتحسينها للاستخدام في Excel
    
    Args:
        image_url: رابط الصورة في Cloudinary
        max_size: الحد الأقصى لحجم الصورة
    
    Returns:
        BytesIO: الصورة المحسنة أو None في حالة الخطأ
    """
    try:
        # تحميل الصورة
        response = requests.get(image_url, timeout=10)
        response.raise_for_status()
        
        # فتح الصورة وتحسينها
        img = PILImage.open(BytesIO(response.content))
        
        # تحويل إلى RGB إذا لزم الأمر
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGB')
        
        # تغيير الحجم مع الحفاظ على النسبة
        img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
        
        # حفظ في BytesIO بجودة عالية
        img_buffer = BytesIO()
        img.save(img_buffer, format='JPEG', quality=95, optimize=True)
        img_buffer.seek(0)
        
        return img_buffer
        
    except Exception as e:
        print(f"خطأ في تحميل الصورة {image_url}: {e}")
        return None

def create_enhanced_excel_with_images(data_entries, filename):
    """
    إنشاء ملف Excel محسن مع الصور والتنسيق
    
    Args:
        data_entries: بيانات الإدخالات
        filename: اسم الملف
    
    Returns:
        str: مسار الملف المؤقت أو None في حالة الخطأ
    """
    try:
        # إنشاء workbook جديد
        wb = Workbook()
        ws = wb.active
        ws.title = "POP Materials Report"
        
        # تعريف الألوان والأنماط
        colors = {
            'header': 'FF366092',
            'alt_row': 'FFF2F2F2',
            'border': 'FF000000',
            'text': 'FF000000',
            'white': 'FFFFFFFF'
        }
        
        # تعريف الخطوط
        header_font = Font(name='Arial', size=12, bold=True, color=colors['white'])
        data_font = Font(name='Arial', size=10, color=colors['text'])
        
        # تعريف الحدود
        thin_border = Border(
            left=Side(style='thin', color=colors['border']),
            right=Side(style='thin', color=colors['border']),
            top=Side(style='thin', color=colors['border']),
            bottom=Side(style='thin', color=colors['border'])
        )
        
        # العناوين بالإنجليزية
        headers = [
            'ID', 'Employee Name', 'Employee Code', 'Branch', 'Shop Code', 
            'Model', 'Display Type', 'Selected Materials', 'Missing Materials', 
            'Images Count', 'Date', 'Image Preview'
        ]
        
        # إضافة العناوين
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # تعيين عرض الأعمدة (مع مساحة إضافية للصور المتعددة)
        column_widths = {
            'A': 8,   # ID
            'B': 20,  # Employee Name
            'C': 15,  # Employee Code
            'D': 25,  # Branch
            'E': 15,  # Shop Code
            'F': 30,  # Model
            'G': 20,  # Display Type
            'H': 40,  # Selected Materials
            'I': 40,  # Missing Materials
            'J': 12,  # Images Count
            'K': 18,  # Date
            'L': 20,  # Image Preview 1
            'M': 20,  # Image Preview 2
            'N': 20   # Image Preview 3
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # إضافة البيانات
        current_row = 2
        
        for entry in data_entries:
            # تحديد لون الصف (متناوب)
            row_fill = PatternFill(
                start_color=colors['alt_row'] if current_row % 2 == 0 else colors['white'],
                end_color=colors['alt_row'] if current_row % 2 == 0 else colors['white'],
                fill_type='solid'
            )
            
            # البيانات الأساسية
            entry_data = [
                entry[0],  # ID
                entry[1],  # Employee Name
                entry[2],  # Employee Code
                entry[3],  # Branch
                entry[4] if entry[4] else 'N/A',  # Shop Code
                entry[5],  # Model
                entry[6],  # Display Type
                entry[7] if entry[7] else 'None',  # Selected Materials
                entry[8] if entry[8] else 'None',  # Missing Materials
                len(entry[9].split(',')) if entry[9] else 0,  # Images Count
                entry[10]  # التاريخ
            ]
            
            # إضافة البيانات إلى الخلايا
            for col, value in enumerate(entry_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right' if col <= 6 else 'right', 
                                         vertical='center', wrap_text=True)
            
            # معالجة الصور
            images_data = entry[9] if entry[9] else ''
            if images_data:
                image_urls = [url.strip() for url in images_data.split(',') if url.strip()]
                
                # تعيين ارتفاع الصف للصور الأكبر
                ws.row_dimensions[current_row].height = 150
                
                # إضافة الصور (حتى 3 صور لكل صف)
                images_added = 0
                for i, image_url in enumerate(image_urls[:3]):  # حتى 3 صور لكل صف
                    img_buffer = None
                    
                    if image_url.startswith('http'):  # Cloudinary URL
                        img_buffer = download_image_from_cloudinary(image_url)
                    else:  # Local file
                        try:
                            # محاولة قراءة الملف المحلي
                            local_path = os.path.join('static/uploads', image_url)
                            if os.path.exists(local_path):
                                with open(local_path, 'rb') as f:
                                    img_data = f.read()
                                
                                # تحسين الصورة
                                img = PILImage.open(BytesIO(img_data))
                                if img.mode in ('RGBA', 'LA', 'P'):
                                    img = img.convert('RGB')
                                
                                img.thumbnail((200, 200), PILImage.Resampling.LANCZOS)
                                
                                img_buffer = BytesIO()
                                img.save(img_buffer, format='JPEG', quality=95, optimize=True)
                                img_buffer.seek(0)
                        except Exception as e:
                            print(f"Error processing local image {image_url}: {e}")
                    
                    if img_buffer:
                        try:
                            # إنشاء صورة Excel بحجم أكبر ووضوح أفضل
                            excel_img = ExcelImage(img_buffer)
                            excel_img.width = 120
                            excel_img.height = 120
                            
                            # تحديد موقع الصورة (توزيع الصور أفقياً)
                            col_letter = get_column_letter(12 + i)  # العمود L, M, N للصور المتعددة
                            anchor_cell = f"{col_letter}{current_row}"
                            
                            # إضافة الصورة
                            ws.add_image(excel_img, anchor_cell)
                            images_added += 1
                            
                        except Exception as e:
                            print(f"Error adding image to Excel: {e}")
                
                # إضافة نص في خلية الصور
                if images_added > 0:
                    img_text = f"{images_added} of {len(image_urls)} images"
                else:
                    img_text = f"0 of {len(image_urls)} images (failed to load)"
                
                img_cell = ws.cell(row=current_row, column=12, value=img_text)
                img_cell.font = data_font
                img_cell.fill = row_fill
                img_cell.border = thin_border
                img_cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # لا توجد صور
                img_cell = ws.cell(row=current_row, column=12, value="No images")
                img_cell.font = data_font
                img_cell.fill = row_fill
                img_cell.border = thin_border
                img_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
        
        # إضافة ملخص في النهاية
        summary_row = current_row + 2
        
        # عنوان الملخص
        summary_cell = ws.cell(row=summary_row, column=1, value="Report Summary")
        summary_cell.font = Font(name='Arial', size=14, bold=True, color=colors['white'])
        summary_cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
        ws.merge_cells(f'A{summary_row}:L{summary_row}')
        summary_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # بيانات الملخص
        total_entries = len(data_entries)
        total_images = sum(len(entry[9].split(',')) if entry[9] else 0 for entry in data_entries)
        unique_employees = len(set(entry[2] for entry in data_entries))
        unique_branches = len(set(entry[3] for entry in data_entries))
        
        summary_data = [
            f"Total Entries: {total_entries}",
            f"Total Images: {total_images}",
            f"Unique Employees: {unique_employees}",
            f"Unique Branches: {unique_branches}",
            f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]
        
        for i, summary_text in enumerate(summary_data):
            cell = ws.cell(row=summary_row + 1 + i, column=1, value=summary_text)
            cell.font = Font(name='Arial', size=10, bold=True)
            ws.merge_cells(f'A{summary_row + 1 + i}:L{summary_row + 1 + i}')
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # حفظ الملف
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, filename)
        wb.save(temp_path)
        
        return temp_path
        
    except Exception as e:
        print(f"خطأ في إنشاء ملف Excel: {e}")
        return None

def export_enhanced_excel_with_cloudinary(data_entries):
    """
    تصدير Excel محسن مع رفع إلى Cloudinary
    
    Args:
        data_entries: بيانات الإدخالات
    
    Returns:
        dict: نتيجة العملية مع رابط التحميل أو الملف المحلي
    """
    try:
        # إنشاء اسم الملف
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'pop_materials_report_enhanced_{timestamp}.xlsx'
        
        # إنشاء ملف Excel محسن
        temp_path = create_enhanced_excel_with_images(data_entries, filename)
        
        if not temp_path:
            return {'success': False, 'error': 'فشل في إنشاء ملف Excel المحسن'}
        
        # محاولة رفع إلى Cloudinary
        if is_cloudinary_configured():
            result = upload_excel_to_cloudinary(temp_path, filename, "reports/enhanced")
            
            if result['success']:
                cleanup_temp_file(temp_path)
                return {
                    'success': True,
                    'method': 'cloudinary',
                    'url': result['url'],
                    'filename': filename,
                    'message': 'تم إنشاء التقرير ورفعه إلى السحابة بنجاح'
                }
        
        # Fallback: إرجاع الملف المحلي
        try:
            with open(temp_path, 'rb') as f:
                file_data = f.read()
            
            cleanup_temp_file(temp_path)
        except FileNotFoundError:
            return {'success': False, 'error': f'الملف المؤقت مفقود: {temp_path}'}
        
        return {
            'success': True,
            'method': 'local',
            'data': file_data,
            'filename': filename,
            'message': 'تم إنشاء التقرير بنجاح'
        }
        
    except Exception as e:
        return {'success': False, 'error': f'خطأ في تصدير Excel: {str(e)}'}

def create_simple_excel_with_formatting(data_entries, filename):
    """
    إنشاء ملف Excel بسيط مع تنسيق جيد (بدون صور)
    
    Args:
        data_entries: بيانات الإدخالات
        filename: اسم الملف
    
    Returns:
        str: مسار الملف المؤقت
    """
    try:
        # تحويل البيانات إلى DataFrame
        df_data = []
        for entry in data_entries:
            df_data.append({
                'ID': entry[0],
                'Employee Name': entry[1],
                'Employee Code': entry[2],
                'Branch': entry[3],
                'Shop Code': entry[4] if entry[4] else 'N/A',
                'Model': entry[5],
                'Display Type': entry[6],
                'Selected Materials': entry[7] if entry[7] else 'None',
                'Missing Materials': entry[8] if entry[8] else 'None',
                'Image Links': entry[9] if entry[9] else 'No images',
                'Date': entry[10]
            })
        
        df = pd.DataFrame(df_data)
        
        # حفظ مع تنسيق
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, filename)
        
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='POP Materials Data', index=False)
            
            # تحسين التنسيق
            worksheet = writer.sheets['POP Materials Data']
            
            # تنسيق العناوين
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # تعديل عرض الأعمدة
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return temp_path
        
    except Exception as e:
        print(f"خطأ في إنشاء ملف Excel البسيط: {e}")
        return None